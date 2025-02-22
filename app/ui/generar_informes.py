from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton, QMessageBox, QFileDialog, QDateEdit
)
from PyQt5.QtCore import QDate
from datetime import datetime

class GenerarInformes(QWidget):
    def __init__(self, controller):
        super().__init__()
        self.controller = controller
        self.setWindowTitle("Generar Informes")
        self.resize(400, 250)

        # Layout principal
        layout = QVBoxLayout()

        # Etiqueta de título
        title_label = QLabel("Generación de Informes")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold;")
        layout.addWidget(title_label)

        # Filtros de fechas
        layout.addWidget(QLabel("Rango de Fechas:"))
        layout.addWidget(QLabel("Desde:"))
        self.fecha_inicio = QDateEdit(QDate.currentDate())
        self.fecha_inicio.setCalendarPopup(True)
        layout.addWidget(self.fecha_inicio)

        layout.addWidget(QLabel("Hasta:"))
        self.fecha_fin = QDateEdit(QDate.currentDate())
        self.fecha_fin.setCalendarPopup(True)
        layout.addWidget(self.fecha_fin)

        # Botón para generar informe
        btn_generar = QPushButton("Generar Informe en Excel")
        btn_generar.clicked.connect(self.generar_informe)
        layout.addWidget(btn_generar)

        self.setLayout(layout)

    def generar_informe(self):
        """Genera un informe en formato Excel válido."""
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_defecto = f"informe_{fecha_actual}.xlsx"

        # Seleccionar dónde guardar el archivo
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar Informe", nombre_defecto, "Archivos Excel (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Informe de Consumos"

            # Encabezados del informe
            columnas = [
                "ID", "Placa", "Cliente", "Canal", "Servicio Propio", "Producto",
                "Repuesto", "Servicio de Terceros", "Total Factura", "Comisión Canal", "Fecha de Factura"
            ]
            sheet.append(columnas)

            # Obtener las fechas seleccionadas
            fecha_inicio = self.fecha_inicio.date().toString("yyyy-MM-dd")
            fecha_fin = self.fecha_fin.date().toString("yyyy-MM-dd")

            # Obtener datos desde el controlador
            datos = self.controller.obtener_informe(fecha_inicio, fecha_fin)

            # Validar si hay datos
            if not datos:
                QMessageBox.warning(self, "Sin Datos", "No hay registros para el rango de fechas seleccionado.")
                return

            # Insertar datos en el archivo Excel
            for row in datos:
                sheet.append(row)

            # Aplicar formato al archivo Excel
            aplicar_formato_excel(sheet)

            # Asegurarse de que la extensión sea .xlsx
            if not file_path.endswith(".xlsx"):
                file_path += ".xlsx"

            workbook.save(file_path)
            QMessageBox.information(self, "Éxito", f"Informe guardado en:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar el informe: {e}")

# Función para aplicar formato al archivo Excel
def aplicar_formato_excel(sheet):
    """Aplica formato básico al archivo Excel."""
    for col in range(1, sheet.max_column + 1):
        letra_col = get_column_letter(col)
        sheet.column_dimensions[letra_col].width = 15  # Ancho de columnas
        cell = sheet[f"{letra_col}1"]
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
